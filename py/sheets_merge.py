import openpyxl

wb = openpyxl.load_workbook("data_2106.xlsx")

# 1つ目のシートか判定
is_first_sheet = True

# リストに全シートから読み取ったデータを集約
row_list = []

# ブック中の全シートを処理
for ws in wb.worksheets:

    if is_first_sheet:
        start_row = 1
    else:
        start_row = 4

    for row in ws.iter_rows(min_row=start_row):
        # 空行で読み込み終了
        if row[0].row > 3 and row[0].value is None:
            break
        value_list = []
        for c in row:
            value_list.append(c.value)
        row_list.append(value_list)

    is_first_sheet = False

# データ転記先のシートを新規作成
ws_new = wb.create_sheet(title="第1四半期売上")

# 書き込み時の行番号
row_num = 1

# 新しいシートに1行ずつデータを書き込む
for row in row_list:
    # 1行分のデータを書き込む
    ws_new.append(row)

    # データ部分のA列に日付の表示形式を設定し、F列の数式を書き換え
    if row_num > 3:
        ws_new.cell(row_num, 1).number_format = "yyyy/m/d"
        ws_new.cell(row_num, 6).value = "=D" + str(row_num) + "*E" + str(row_num)
    row_num = row_num + 1

# 別名でブック保存
wb.save("data_1Q.xlsx")
